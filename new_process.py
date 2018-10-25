# -*-coding:utf-8-*-
from openpyxl import load_workbook
from openpyxl import Workbook
# 加载Excel工作表以及获取表名
wb1 = load_workbook(r"C:\Users\wujian\Desktop\work\source.xlsx")
sheet1 = wb1.get_sheet_by_name("sheet1")
# 按照“-”进行字符串的截取
def split(s):
    if '-' in s:
        index=s.rindex("-")
        s=s[0:index]
    if "-" in s:
       index=s.rindex("-")
       s=s[0:index]
    return s

# 把所有的井号单独取出来
well_set1=[]
for x in sheet1['B']:
    if x.value!='井号':
        well_set1.append(x.value)

# 读取原始数据的行数，不包括原始excel的第一行
s_max_row=len(well_set1)
# 对井号去重排序
well_set=list(set(well_set1))
well_set.sort()
# 发现有多少口井---及结果新表中会有多少行
max_row=len(well_set)


# for x in well_set:
#     print(x)

# 把所有的成本属性单独取出来，
attr_set1=[]
for x in sheet1['D']:
    if x.value!='成本科目+模板':
        v=split(x.value)
        attr_set1.append(v)

# 对成本属性集合去重、排序
attr_set=list(set(attr_set1))
attr_set.sort()
# 读取到有多少个属性值，---即结果的新表中有多少列
max_col=len(attr_set)

# for x in attr_set:
#     print(x)

# 设计一个列表的嵌套，all_data=[["井号"，"成本属性","成本金额"],[]]用于存放原始的数据信息，以便在内存中操作数据
all_data=[]
for i in range(2,s_max_row+2):
    item=[]
    item.append(sheet1['B'+str(i)].value)# cell 单元的定位是‘B2’样式，不是‘2B’
    x=split(sheet1['D'+str(i)].value)
    item.append(x)
    item.append(sheet1['H'+str(i)].value)
    all_data.append(item)
    del item

all_data.sort()
print(len(all_data))
for x in all_data:
    print(x)

# 设计一个嵌套字典结构用于处理数据 --将每口井的属性和成本汇集在一起，井号作为key，属性与金额作为value
# data_dict={'井号1':{'成本属性1':成本金额1,'成本属性2':成本金额2,,,,},
#           '井号2':{'成本属性1':成本金额1,'成本属性2':成本金额2,,,,},
#           ...}
data_dict={}
j=0
for i in range(s_max_row):
    key=well_set[j]
    attr={}   # 每口井的属性
    attr[all_data[i][1]]=all_data[i][2]
    print(attr)
    print(all_data[i])
    if key==all_data[i][0]:
        # 如果字典中已经存在某井号的key
        if key in data_dict.keys():
            data_dict[key].update(attr)
        else:
            data_dict[key]=attr
    else:
        j+=1
        data_dict[well_set[j]]=attr
    del attr


# for x in data_dict:
#     print(x, data_dict[x])

# 处理后的数据已经准备好了，现在可以把结果数据写入到新的excel表格了
# 创建一个新的Excel工作表和sheet表单
wb2 = Workbook()
sheet2 = wb2.active
sheet2.title="sheet2"
alphabet=['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
        'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT',
        'AU','AV','AW','AX','AY','AZ','BA']
sheet2['A1']='井号'
# 先把所有的成本属性写入到第一行中
for i in range(max_col):
    index='%s%d'%(alphabet[i],1)
    sheet2[index].value=attr_set[i]

# 将每一行的数据写入
for i in range(max_row):
    # 确立每一行的井号
    well=well_set[i]
    sheet2['A'+str(i+2)].value=well
    # 对每一行的每个cell进行遍历，把相应的成本数据写入进去
    for j in range(52):
        # 确立每一个cell的位置
        index='%s%d'%(alphabet[j],i+2)
        key_attr=attr_set[j]
        if key_attr in data_dict[well].keys():
            sheet2[index].value=data_dict[well][key_attr]

# 将最终的结果保存在excel中
wb2.save(r"C:\Users\wujian\Desktop\work\source2.xlsx")

